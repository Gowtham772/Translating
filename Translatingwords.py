import openpyxl
import re
s=open("C:\\Users\\Intel\\Desktop\\exeter\\t8.shakespeare.txt","r")
toread=s.read()
res=toread.lower()
wb=openpyxl.load_workbook("C:\\Users\\Intel\\Desktop\\exeter\\french_dictionary.xlsx")
sheets=wb.sheetnames
sh1=wb['Sheet1']
words_dict={}
print('English word , French word , Frequency')
for i in range(1,1000):
    a=sh1.cell(i,1).value
    c=str(a)
    if c in res:
        b=sh1.cell(i,2).value
        print(c,',',b,',',res.count(c))
        replace={c:b}
        for key, value in replace.items():
            res=res.replace(key, value)
print(res)
